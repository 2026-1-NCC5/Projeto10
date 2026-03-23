from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SCRIPT_DIR = Path(__file__).resolve().parent
CSV_PATH = SCRIPT_DIR / "pixel_data.csv"
EXCEL_PATH = SCRIPT_DIR / "representacao_matricial.xlsx"


def create_excel(csv_path, excel_path):
    df = pd.read_csv(csv_path).astype(int)

    wb = Workbook()

    ws_data = wb.active
    ws_data.title = "Dados dos Pixels"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="00AB72", end_color="00AB72", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    headers = list(df.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for row_idx, row in df.iterrows():
        for col_idx, value in enumerate(row, 1):
            cell = ws_data.cell(row=row_idx + 2, column=col_idx, value=int(value))
            cell.alignment = center

    for col_idx in range(1, len(headers) + 1):
        ws_data.column_dimensions[get_column_letter(col_idx)].width = 8

    ws_data.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(df) + 1}"

    ws_img = wb.create_sheet("Imagem Reconstruída")

    height = df["Y"].max() + 1
    width = df["X"].max() + 1

    pixel_size = 3.0
    for col_idx in range(1, width + 1):
        ws_img.column_dimensions[get_column_letter(col_idx)].width = pixel_size

    for row_idx in range(1, height + 1):
        ws_img.row_dimensions[row_idx].height = pixel_size * 5.5

    for _, row in df.iterrows():
        y, x, r, g, b = int(row["Y"]), int(row["X"]), int(row["R"]), int(row["G"]), int(row["B"])
        color = f"{r:02X}{g:02X}{b:02X}"
        cell = ws_img.cell(row=y + 1, column=x + 1)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    wb.save(excel_path)
    return excel_path


def main():
    if not CSV_PATH.exists():
        print(f"CSV não encontrado: {CSV_PATH}")
        print("Execute primeiro o main.py ou o notebook para gerar o pixel_data.csv")
        return

    path = create_excel(CSV_PATH, EXCEL_PATH)
    print(f"Excel gerado: {path.name}")
    print(f"  - Aba 'Dados dos Pixels': 2.500 linhas com Y, X, R, G, B")
    print(f"  - Aba 'Imagem Reconstruída': cada célula pintada com a cor RGB do pixel")


if __name__ == "__main__":
    main()
